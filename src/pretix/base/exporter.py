import io
import tempfile
from collections import OrderedDict, namedtuple
from decimal import Decimal
from typing import Tuple

from defusedcsv import csv
from django import forms
from django.db.models import QuerySet
from django.utils.formats import localize
from django.utils.translation import gettext, gettext_lazy as _
from openpyxl import Workbook
from openpyxl.cell.cell import KNOWN_TYPES

from pretix.base.models import Event


class BaseExporter:
    """
    This is the base class for all data exporters
    """

    def __init__(self, event, progress_callback=lambda v: None):
        self.event = event
        self.progress_callback = progress_callback
        self.is_multievent = isinstance(event, QuerySet)
        if isinstance(event, QuerySet):
            self.events = event
            self.event = None
        else:
            self.events = Event.objects.filter(pk=event.pk)

    def __str__(self):
        return self.identifier

    @property
    def verbose_name(self) -> str:
        """
        A human-readable name for this exporter. This should be short but
        self-explaining. Good examples include 'JSON' or 'Microsoft Excel'.
        """
        raise NotImplementedError()  # NOQA

    @property
    def identifier(self) -> str:
        """
        A short and unique identifier for this exporter.
        This should only contain lowercase letters and in most
        cases will be the same as your package name.
        """
        raise NotImplementedError()  # NOQA

    @property
    def export_form_fields(self) -> dict:
        """
        When the event's administrator visits the export page, this method
        is called to return the configuration fields available.

        It should therefore return a dictionary where the keys should be field names and
        the values should be corresponding Django form fields.

        We suggest that you return an ``OrderedDict`` object instead of a dictionary.
        Your implementation could look like this::

            @property
            def export_form_fields(self):
                return OrderedDict(
                    [
                        ('tab_width',
                         forms.IntegerField(
                             label=_('Tab width'),
                             default=4
                         ))
                    ]
                )
        """
        return {}

    def render(self, form_data: dict) -> Tuple[str, str, bytes]:
        """
        Render the exported file and return a tuple consisting of a filename, a file type
        and file content.

        :type form_data: dict
        :param form_data: The form data of the export details form
        :param output_file: You can optionally accept a parameter that will be given a file handle to write the
                            output to. In this case, you can return None instead of the file content.

        Note: If you use a ``ModelChoiceField`` (or a ``ModelMultipleChoiceField``), the
        ``form_data`` will not contain the model instance but only it's primary key (or
        a list of primary keys) for reasons of internal serialization when using background
        tasks.
        """
        raise NotImplementedError()  # NOQA


class ListExporter(BaseExporter):
    ProgressSetTotal = namedtuple('ProgressSetTotal', 'total')

    @property
    def export_form_fields(self) -> dict:
        ff = OrderedDict(
            [
                ('_format',
                 forms.ChoiceField(
                     label=_('Export format'),
                     choices=(
                         ('xlsx', _('Excel (.xlsx)')),
                         ('default', _('CSV (with commas)')),
                         ('csv-excel', _('CSV (Excel-style)')),
                         ('semicolon', _('CSV (with semicolons)')),
                     ),
                 )),
            ]
        )
        ff.update(self.additional_form_fields)
        return ff

    @property
    def additional_form_fields(self) -> dict:
        return {}

    def iterate_list(self, form_data):
        raise NotImplementedError()  # noqa

    def get_filename(self):
        return 'export'

    def _render_csv(self, form_data, output_file=None, **kwargs):
        if output_file:
            writer = csv.writer(output_file, **kwargs)
            total = 0
            counter = 0
            for line in self.iterate_list(form_data):
                if isinstance(line, self.ProgressSetTotal):
                    total = line.total
                    continue
                line = [
                    localize(f) if isinstance(f, Decimal) else f
                    for f in line
                ]
                if total:
                    counter += 1
                    if counter % max(10, total // 100) == 0:
                        self.progress_callback(counter / total * 100)
                writer.writerow(line)
            return self.get_filename() + '.csv', 'text/csv', None
        else:
            output = io.StringIO()
            writer = csv.writer(output, **kwargs)
            total = 0
            counter = 0
            for line in self.iterate_list(form_data):
                if isinstance(line, self.ProgressSetTotal):
                    total = line.total
                    continue
                line = [
                    localize(f) if isinstance(f, Decimal) else f
                    for f in line
                ]
                if total:
                    counter += 1
                    if counter % max(10, total // 100) == 0:
                        self.progress_callback(counter / total * 100)
                writer.writerow(line)
            return self.get_filename() + '.csv', 'text/csv', output.getvalue().encode("utf-8")

    def _render_xlsx(self, form_data, output_file=None):
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        try:
            ws.title = str(self.verbose_name)
        except:
            pass
        total = 0
        counter = 0
        for i, line in enumerate(self.iterate_list(form_data)):
            if isinstance(line, self.ProgressSetTotal):
                total = line.total
                continue
            ws.append([
                str(val) if not isinstance(val, KNOWN_TYPES) else val
                for val in line
            ])
            if total:
                counter += 1
                if counter % max(10, total // 100) == 0:
                    self.progress_callback(counter / total * 100)

        if output_file:
            wb.save(output_file)
            return self.get_filename() + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', None
        else:
            with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
                wb.save(f.name)
                f.seek(0)
                return self.get_filename() + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f.read()

    def render(self, form_data: dict, output_file=None) -> Tuple[str, str, bytes]:
        if form_data.get('_format') == 'xlsx':
            return self._render_xlsx(form_data, output_file=output_file)
        elif form_data.get('_format') == 'default':
            return self._render_csv(form_data, quoting=csv.QUOTE_NONNUMERIC, delimiter=',', output_file=output_file)
        elif form_data.get('_format') == 'csv-excel':
            return self._render_csv(form_data, dialect='excel', output_file=output_file)
        elif form_data.get('_format') == 'semicolon':
            return self._render_csv(form_data, dialect='excel', delimiter=';', output_file=output_file)


class MultiSheetListExporter(ListExporter):

    @property
    def sheets(self):
        raise NotImplementedError()

    @property
    def export_form_fields(self) -> dict:
        choices = [
            ('xlsx', _('Combined Excel (.xlsx)')),
        ]
        for s, l in self.sheets:
            choices += [
                (s + ':default', str(l) + ' – ' + gettext('CSV (with commas)')),
                (s + ':excel', str(l) + ' – ' + gettext('CSV (Excel-style)')),
                (s + ':semicolon', str(l) + ' – ' + gettext('CSV (with semicolons)')),
            ]
        ff = OrderedDict(
            [
                ('_format',
                 forms.ChoiceField(
                     label=_('Export format'),
                     choices=choices,
                 )),
            ]
        )
        ff.update(self.additional_form_fields)
        return ff

    def iterate_list(self, form_data):
        pass

    def iterate_sheet(self, form_data, sheet):
        raise NotImplementedError()  # noqa

    def _render_sheet_csv(self, form_data, sheet, output_file=None, **kwargs):
        total = 0
        counter = 0
        if output_file:
            writer = csv.writer(output_file, **kwargs)
            for line in self.iterate_sheet(form_data, sheet):
                if isinstance(line, self.ProgressSetTotal):
                    total = line.total
                    continue
                line = [
                    localize(f) if isinstance(f, Decimal) else f
                    for f in line
                ]
                writer.writerow(line)
                if total:
                    counter += 1
                    if counter % max(10, total // 100) == 0:
                        self.progress_callback(counter / total * 100)
            return self.get_filename() + '.csv', 'text/csv', None
        else:
            output = io.StringIO()
            writer = csv.writer(output, **kwargs)
            for line in self.iterate_sheet(form_data, sheet):
                if isinstance(line, self.ProgressSetTotal):
                    total = line.total
                    continue
                line = [
                    localize(f) if isinstance(f, Decimal) else f
                    for f in line
                ]
                writer.writerow(line)
                if total:
                    counter += 1
                    if counter % max(10, total // 100) == 0:
                        self.progress_callback(counter / total * 100)
            return self.get_filename() + '.csv', 'text/csv', output.getvalue().encode("utf-8")

    def _render_xlsx(self, form_data, output_file=None):
        wb = Workbook(write_only=True)
        n_sheets = len(self.sheets)
        for i_sheet, (s, l) in enumerate(self.sheets):
            ws = wb.create_sheet(str(l))
            total = 0
            counter = 0
            for i, line in enumerate(self.iterate_sheet(form_data, sheet=s)):
                if isinstance(line, self.ProgressSetTotal):
                    total = line.total
                    continue
                ws.append([
                    str(val) if not isinstance(val, KNOWN_TYPES) else val
                    for val in line
                ])
                if total:
                    counter += 1
                    if counter % max(10, total // 100) == 0:
                        self.progress_callback(counter / total * 100 / n_sheets + 100 / n_sheets * i_sheet)

        if output_file:
            wb.save(output_file)
            return self.get_filename() + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', None
        else:
            with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
                wb.save(f.name)
                f.seek(0)
                return self.get_filename() + '.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f.read()

    def render(self, form_data: dict, output_file=None) -> Tuple[str, str, bytes]:
        if form_data.get('_format') == 'xlsx':
            return self._render_xlsx(form_data, output_file=output_file)
        elif ':' in form_data.get('_format'):
            sheet, f = form_data.get('_format').split(':')
            if f == 'default':
                return self._render_sheet_csv(form_data, sheet, quoting=csv.QUOTE_NONNUMERIC, delimiter=',',
                                              output_file=output_file)
            elif f == 'excel':
                return self._render_sheet_csv(form_data, sheet, dialect='excel', output_file=output_file)
            elif f == 'semicolon':
                return self._render_sheet_csv(form_data, sheet, dialect='excel', delimiter=';', output_file=output_file)
