#
# This file is part of pretix (Community Edition).
#
# Copyright (C) 2014-2020 Raphael Michel and contributors
# Copyright (C) 2020-2021 rami.io GmbH and contributors
#
# This program is free software: you can redistribute it and/or modify it under the terms of the GNU Affero General
# Public License as published by the Free Software Foundation in version 3 of the License.
#
# ADDITIONAL TERMS APPLY: Pursuant to Section 7 of the GNU Affero General Public License, additional terms are
# applicable granting you additional permissions and placing additional restrictions on your usage of this software.
# Please refer to the pretix LICENSE file to obtain the full terms applicable to this work. If you did not receive
# this file, see <https://pretix.eu/about/en/license>.
#
# This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even the implied
# warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU Affero General Public License for more
# details.
#
# You should have received a copy of the GNU Affero General Public License along with this program.  If not, see
# <https://www.gnu.org/licenses/>.
#

# This file is based on an earlier version of pretix which was released under the Apache License 2.0. The full text of
# the Apache License 2.0 can be obtained at <http://www.apache.org/licenses/LICENSE-2.0>.
#
# This file may have since been changed and any changes are released under the terms of AGPLv3 as described above. A
# full history of changes and contributors is available at <https://github.com/pretix/pretix>.
#
# This file contains Apache-licensed contributions copyrighted by: Benjamin Hättasch, Tobias Kunze
#
# Unless required by applicable law or agreed to in writing, software distributed under the Apache License 2.0 is
# distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
# License for the specific language governing permissions and limitations under the License.

from collections import OrderedDict
from decimal import Decimal
from zoneinfo import ZoneInfo

from django import forms
from django.db.models import (
    Case, CharField, Count, DateTimeField, F, IntegerField, Max, Min, OuterRef,
    Q, Subquery, Sum, When,
)
from django.db.models.functions import Coalesce
from django.dispatch import receiver
from django.utils.formats import date_format
from django.utils.functional import cached_property
from django.utils.timezone import get_current_timezone, now
from django.utils.translation import (
    gettext as _, gettext_lazy, pgettext, pgettext_lazy,
)
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from pretix.base.models import (
    GiftCard, GiftCardTransaction, Invoice, InvoiceAddress, Order,
    OrderPosition, Question,
)
from pretix.base.models.orders import (
    OrderFee, OrderPayment, OrderRefund, Transaction,
)
from pretix.base.services.quotas import QuotaAvailability
from pretix.base.settings import PERSON_NAME_SCHEMES, get_name_parts_localized

from ...control.forms.filter import get_all_payment_providers
from ...helpers import GroupConcat
from ...helpers.iter import chunked_iterable
from ...helpers.safe_openpyxl import remove_invalid_excel_chars
from ..exporter import (
    ListExporter, MultiSheetListExporter, OrganizerLevelExportMixin,
)
from ..forms.widgets import SplitDateTimePickerWidget
from ..signals import (
    register_data_exporters, register_multievent_data_exporters,
)
from ..timeframes import (
    DateFrameField,
    resolve_timeframe_to_datetime_start_inclusive_end_exclusive,
)


class OrderListExporter(MultiSheetListExporter):
    identifier = 'orderlist'
    verbose_name = gettext_lazy('Order data')
    category = pgettext_lazy('export_category', 'Order data')
    description = gettext_lazy('Download a spreadsheet of all orders. The spreadsheet will include three sheets, one '
                               'with a line for every order, one with a line for every order position, and one with '
                               'a line for every additional fee charged in an order.')
    featured = True

    @cached_property
    def providers(self):
        return dict(get_all_payment_providers())

    @property
    def sheets(self):
        return (
            ('orders', _('Orders')),
            ('positions', _('Order positions')),
            ('fees', _('Order fees')),
        )

    @property
    def additional_form_fields(self):
        d = [
            ('paid_only',
             forms.BooleanField(
                 label=_('Only paid orders'),
                 initial=True,
                 required=False
             )),
            ('include_payment_amounts',
             forms.BooleanField(
                 label=_('Include payment amounts'),
                 initial=False,
                 required=False
             )),
            ('group_multiple_choice',
             forms.BooleanField(
                 label=_('Show multiple choice answers grouped in one column'),
                 initial=False,
                 required=False
             )),
            ('date_range',
             DateFrameField(
                 label=_('Date range'),
                 include_future_frames=False,
                 required=False,
                 help_text=_('Only include orders created within this date range.')
             )),
            ('event_date_range',
             DateFrameField(
                 label=_('Event date'),
                 include_future_frames=True,
                 required=False,
                 help_text=_('Only include orders including at least one ticket for a date in this range. '
                             'Will also include other dates in case of mixed orders!')
             )),
        ]
        d = OrderedDict(d)
        if not self.is_multievent and not self.event.has_subevents:
            del d['event_date_range']
        return d

    def _get_all_payment_methods(self, qs):
        pps = dict(get_all_payment_providers())
        return sorted([(pp, pps[pp]) for pp in set(
            OrderPayment.objects.exclude(provider='free').filter(order__event__in=self.events).values_list(
                'provider', flat=True
            ).distinct()
        )], key=lambda pp: pp[0])

    def _get_all_tax_rates(self, qs):
        tax_rates = set(
            a for a
            in OrderFee.objects.filter(
                order__event__in=self.events
            ).values_list('tax_rate', flat=True).distinct().order_by()
        )
        tax_rates |= set(
            a for a
            in OrderPosition.objects.filter(
                order__event__in=self.events
            ).values_list('tax_rate', flat=True).distinct().order_by()
        )
        tax_rates = sorted(tax_rates)
        return tax_rates

    def iterate_sheet(self, form_data, sheet):
        if sheet == 'orders':
            return self.iterate_orders(form_data)
        elif sheet == 'positions':
            return self.iterate_positions(form_data)
        elif sheet == 'fees':
            return self.iterate_fees(form_data)

    @cached_property
    def event_object_cache(self):
        return {e.pk: e for e in self.events}

    def _date_filter(self, qs, form_data, rel):
        annotations = {}
        filters = {}

        if form_data.get('date_range'):
            dt_start, dt_end = resolve_timeframe_to_datetime_start_inclusive_end_exclusive(now(), form_data['date_range'], self.timezone)
            if dt_start:
                filters[f'{rel}datetime__gte'] = dt_start
            if dt_end:
                filters[f'{rel}datetime__lt'] = dt_end

        if form_data.get('event_date_range'):
            dt_start, dt_end = resolve_timeframe_to_datetime_start_inclusive_end_exclusive(now(), form_data['event_date_range'], self.timezone)
            if dt_start:
                annotations['event_date_max'] = Case(
                    When(**{f'{rel}event__has_subevents': True}, then=Max(f'{rel}all_positions__subevent__date_from')),
                    default=F(f'{rel}event__date_from'),
                )
                filters['event_date_max__gte'] = dt_start
            if dt_end:
                annotations['event_date_min'] = Case(
                    When(**{f'{rel}event__has_subevents': True}, then=Min(f'{rel}all_positions__subevent__date_from')),
                    default=F(f'{rel}event__date_from'),
                )
                filters['event_date_min__lt'] = dt_end

        if filters:
            return qs.annotate(**annotations).filter(**filters)
        return qs

    def iterate_orders(self, form_data: dict):
        p_date = OrderPayment.objects.filter(
            order=OuterRef('pk'),
            state__in=(OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED),
            payment_date__isnull=False
        ).values('order').annotate(
            m=Max('payment_date')
        ).values(
            'm'
        ).order_by()
        p_providers = OrderPayment.objects.filter(
            order=OuterRef('pk'),
            state__in=(OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED,
                       OrderPayment.PAYMENT_STATE_PENDING, OrderPayment.PAYMENT_STATE_CREATED),
        ).values('order').annotate(
            m=GroupConcat('provider', delimiter=',')
        ).values(
            'm'
        ).order_by()
        i_numbers = Invoice.objects.filter(
            order=OuterRef('pk'),
        ).values('order').annotate(
            m=GroupConcat('full_invoice_no', delimiter=', ')
        ).values(
            'm'
        ).order_by()

        s = OrderPosition.objects.filter(
            order=OuterRef('pk')
        ).order_by().values('order').annotate(k=Count('id')).values('k')
        qs = Order.objects.filter(event__in=self.events).annotate(
            payment_date=Subquery(p_date, output_field=DateTimeField()),
            payment_providers=Subquery(p_providers, output_field=CharField()),
            invoice_numbers=Subquery(i_numbers, output_field=CharField()),
            pcnt=Subquery(s, output_field=IntegerField())
        ).select_related('invoice_address', 'customer')

        qs = self._date_filter(qs, form_data, rel='')

        if form_data['paid_only']:
            qs = qs.filter(status=Order.STATUS_PAID)
        tax_rates = self._get_all_tax_rates(qs)

        headers = [
            _('Event slug'), _('Order code'), _('Order total'), _('Status'), _('Email'), _('Phone number'),
            _('Order date'), _('Order time'), _('Company'), _('Name'),
        ]
        name_scheme = PERSON_NAME_SCHEMES[self.event.settings.name_scheme] if not self.is_multievent else None
        if name_scheme and len(name_scheme['fields']) > 1:
            for k, label, w in name_scheme['fields']:
                headers.append(label)
        headers += [
            _('Address'), _('ZIP code'), _('City'), _('Country'), pgettext('address', 'State'),
            _('Custom address field'), _('VAT ID'), _('Date of last payment'), _('Fees'), _('Order locale')
        ]

        for tr in tax_rates:
            headers += [
                _('Gross at {rate} % tax').format(rate=tr),
                _('Net at {rate} % tax').format(rate=tr),
                _('Tax value at {rate} % tax').format(rate=tr),
            ]

        headers.append(_('Invoice numbers'))
        headers.append(_('Sales channel'))
        headers.append(_('Requires special attention'))
        headers.append(_('Comment'))
        headers.append(_('Follow-up date'))
        headers.append(_('Positions'))
        headers.append(_('E-mail address verified'))
        headers.append(_('External customer ID'))
        headers.append(_('Payment providers'))
        if form_data.get('include_payment_amounts'):
            payment_methods = self._get_all_payment_methods(qs)
            for id, vn in payment_methods:
                headers.append(_('Paid by {method}').format(method=vn))

        # get meta_data labels from first cached event
        headers += next(iter(self.event_object_cache.values())).meta_data.keys()
        yield headers

        full_fee_sum_cache = {
            o['order__id']: o['grosssum'] for o in
            OrderFee.objects.values('tax_rate', 'order__id').order_by().annotate(grosssum=Sum('value'))
        }
        fee_sum_cache = {
            (o['order__id'], o['tax_rate']): o for o in
            OrderFee.objects.values('tax_rate', 'order__id').order_by().annotate(
                taxsum=Sum('tax_value'), grosssum=Sum('value')
            )
        }
        if form_data.get('include_payment_amounts'):
            payment_sum_cache = {
                (o['order__id'], o['provider']): o['grosssum'] for o in
                OrderPayment.objects.values('provider', 'order__id').order_by().filter(
                    state__in=[OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED]
                ).annotate(
                    grosssum=Sum('amount')
                )
            }
            refund_sum_cache = {
                (o['order__id'], o['provider']): o['grosssum'] for o in
                OrderRefund.objects.values('provider', 'order__id').order_by().filter(
                    state__in=[OrderRefund.REFUND_STATE_DONE, OrderRefund.REFUND_STATE_TRANSIT]
                ).annotate(
                    grosssum=Sum('amount')
                )
            }
        sum_cache = {
            (o['order__id'], o['tax_rate']): o for o in
            OrderPosition.objects.values('tax_rate', 'order__id').order_by().annotate(
                taxsum=Sum('tax_value'), grosssum=Sum('price')
            )
        }

        yield self.ProgressSetTotal(total=qs.count())
        for order in qs.order_by('datetime').iterator():
            tz = ZoneInfo(self.event_object_cache[order.event_id].settings.timezone)

            row = [
                self.event_object_cache[order.event_id].slug,
                order.code,
                order.total,
                order.get_status_display(),
                order.email,
                str(order.phone) if order.phone else '',
                order.datetime.astimezone(tz).strftime('%Y-%m-%d'),
                order.datetime.astimezone(tz).strftime('%H:%M:%S'),
            ]
            try:
                row += [
                    order.invoice_address.company,
                    order.invoice_address.name,
                ]
                if name_scheme and len(name_scheme['fields']) > 1:
                    for k, label, w in name_scheme['fields']:
                        row.append(
                            get_name_parts_localized(order.invoice_address.name_parts, k)
                        )
                row += [
                    order.invoice_address.street,
                    order.invoice_address.zipcode,
                    order.invoice_address.city,
                    order.invoice_address.country if order.invoice_address.country else
                    order.invoice_address.country_old,
                    order.invoice_address.state,
                    order.invoice_address.custom_field,
                    order.invoice_address.vat_id,
                ]
            except InvoiceAddress.DoesNotExist:
                row += [''] * (9 + (len(name_scheme['fields']) if name_scheme and len(name_scheme['fields']) > 1 else 0))

            row += [
                order.payment_date.astimezone(tz).strftime('%Y-%m-%d') if order.payment_date else '',
                full_fee_sum_cache.get(order.id) or Decimal('0.00'),
                order.locale,
            ]

            for tr in tax_rates:
                taxrate_values = sum_cache.get((order.id, tr), {'grosssum': Decimal('0.00'), 'taxsum': Decimal('0.00')})
                fee_taxrate_values = fee_sum_cache.get((order.id, tr),
                                                       {'grosssum': Decimal('0.00'), 'taxsum': Decimal('0.00')})

                row += [
                    taxrate_values['grosssum'] + fee_taxrate_values['grosssum'],
                    (
                        taxrate_values['grosssum'] - taxrate_values['taxsum'] +
                        fee_taxrate_values['grosssum'] - fee_taxrate_values['taxsum']
                    ),
                    taxrate_values['taxsum'] + fee_taxrate_values['taxsum'],
                ]

            row.append(order.invoice_numbers)
            row.append(order.sales_channel)
            row.append(_('Yes') if order.checkin_attention else _('No'))
            row.append(order.comment or "")
            row.append(order.custom_followup_at.strftime("%Y-%m-%d") if order.custom_followup_at else "")
            row.append(order.pcnt)
            row.append(_('Yes') if order.email_known_to_work else _('No'))
            row.append(str(order.customer.external_identifier) if order.customer and order.customer.external_identifier else '')
            row.append(', '.join([
                str(self.providers.get(p, p)) for p in sorted(set((order.payment_providers or '').split(',')))
                if p and p != 'free'
            ]))

            if form_data.get('include_payment_amounts'):
                payment_methods = self._get_all_payment_methods(qs)
                for id, vn in payment_methods:
                    row.append(
                        payment_sum_cache.get((order.id, id), Decimal('0.00')) -
                        refund_sum_cache.get((order.id, id), Decimal('0.00'))
                    )
            row += self.event_object_cache[order.event_id].meta_data.values()
            yield row

    def iterate_fees(self, form_data: dict):
        p_providers = OrderPayment.objects.filter(
            order=OuterRef('order'),
            state__in=(OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED,
                       OrderPayment.PAYMENT_STATE_PENDING, OrderPayment.PAYMENT_STATE_CREATED),
        ).values('order').annotate(
            m=GroupConcat('provider', delimiter=',')
        ).values(
            'm'
        ).order_by()
        qs = OrderFee.all.filter(
            order__event__in=self.events,
        ).annotate(
            payment_providers=Subquery(p_providers, output_field=CharField()),
        ).select_related('order', 'order__invoice_address', 'order__customer', 'tax_rule')
        if form_data['paid_only']:
            qs = qs.filter(order__status=Order.STATUS_PAID, canceled=False)

        qs = self._date_filter(qs, form_data, rel='order__')

        headers = [
            _('Event slug'),
            _('Order code'),
            _('Status'),
            _('Email'),
            _('Phone number'),
            _('Order date'),
            _('Order time'),
            _('Fee type'),
            _('Description'),
            _('Price'),
            _('Tax rate'),
            _('Tax rule'),
            _('Tax value'),
            _('Company'),
            _('Invoice address name'),
        ]
        name_scheme = PERSON_NAME_SCHEMES[self.event.settings.name_scheme] if not self.is_multievent else None
        if name_scheme and len(name_scheme['fields']) > 1:
            for k, label, w in name_scheme['fields']:
                headers.append(_('Invoice address name') + ': ' + str(label))
        headers += [
            _('Address'), _('ZIP code'), _('City'), _('Country'), pgettext('address', 'State'), _('VAT ID'),
        ]

        headers.append(_('External customer ID'))
        headers.append(_('Payment providers'))

        # get meta_data labels from first cached event
        headers += next(iter(self.event_object_cache.values())).meta_data.keys()
        yield headers

        yield self.ProgressSetTotal(total=qs.count())
        for op in qs.order_by('order__datetime').iterator():
            order = op.order
            tz = ZoneInfo(order.event.settings.timezone)
            row = [
                self.event_object_cache[order.event_id].slug,
                order.code,
                _("canceled") if op.canceled else order.get_status_display(),
                order.email,
                str(order.phone) if order.phone else '',
                order.datetime.astimezone(tz).strftime('%Y-%m-%d'),
                order.datetime.astimezone(tz).strftime('%H:%M:%S'),
                op.get_fee_type_display(),
                op.description,
                op.value,
                op.tax_rate,
                str(op.tax_rule) if op.tax_rule else '',
                op.tax_value,
            ]
            try:
                row += [
                    order.invoice_address.company,
                    order.invoice_address.name,
                ]
                if name_scheme and len(name_scheme['fields']) > 1:
                    for k, label, w in name_scheme['fields']:
                        row.append(
                            get_name_parts_localized(order.invoice_address.name_parts, k)
                        )
                row += [
                    order.invoice_address.street,
                    order.invoice_address.zipcode,
                    order.invoice_address.city,
                    order.invoice_address.country if order.invoice_address.country else
                    order.invoice_address.country_old,
                    order.invoice_address.state,
                    order.invoice_address.vat_id,
                ]
            except InvoiceAddress.DoesNotExist:
                row += [''] * (8 + (len(name_scheme['fields']) if name_scheme and len(name_scheme['fields']) > 1 else 0))
            row.append(str(order.customer.external_identifier) if order.customer and order.customer.external_identifier else '')
            row.append(', '.join([
                str(self.providers.get(p, p)) for p in sorted(set((op.payment_providers or '').split(',')))
                if p and p != 'free'
            ]))
            row += self.event_object_cache[order.event_id].meta_data.values()
            yield row

    def iterate_positions(self, form_data: dict):
        p_providers = OrderPayment.objects.filter(
            order=OuterRef('order'),
            state__in=(OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED,
                       OrderPayment.PAYMENT_STATE_PENDING, OrderPayment.PAYMENT_STATE_CREATED),
        ).values('order').annotate(
            m=GroupConcat('provider', delimiter=',')
        ).values(
            'm'
        ).order_by()
        base_qs = OrderPosition.all.filter(
            order__event__in=self.events,
        )
        qs = base_qs.annotate(
            payment_providers=Subquery(p_providers, output_field=CharField()),
        ).select_related(
            'order', 'order__invoice_address', 'order__customer', 'item', 'variation',
            'voucher', 'tax_rule'
        ).prefetch_related(
            'subevent', 'subevent__meta_values',
            'answers', 'answers__question', 'answers__options'
        )
        if form_data['paid_only']:
            qs = qs.filter(order__status=Order.STATUS_PAID, canceled=False)

        qs = self._date_filter(qs, form_data, rel='order__')

        has_subevents = self.events.filter(has_subevents=True).exists()

        headers = [
            _('Event slug'),
            _('Order code'),
            _('Position ID'),
            _('Status'),
            _('Email'),
            _('Phone number'),
            _('Order date'),
            _('Order time'),
        ]
        if has_subevents:
            headers.append(pgettext('subevent', 'Date'))
            headers.append(_('Start date'))
            headers.append(_('End date'))
        headers += [
            _('Product'),
            _('Variation'),
            _('Price'),
            _('Tax rate'),
            _('Tax rule'),
            _('Tax value'),
            _('Attendee name'),
        ]
        name_scheme = PERSON_NAME_SCHEMES[self.event.settings.name_scheme] if not self.is_multievent else None
        if name_scheme and len(name_scheme['fields']) > 1:
            for k, label, w in name_scheme['fields']:
                headers.append(_('Attendee name') + ': ' + str(label))
        headers += [
            _('Attendee email'),
            _('Company'),
            _('Address'),
            _('ZIP code'),
            _('City'),
            _('Country'),
            pgettext('address', 'State'),
            _('Voucher'),
            _('Pseudonymization ID'),
            _('Ticket secret'),
            _('Seat ID'),
            _('Seat name'),
            _('Seat zone'),
            _('Seat row'),
            _('Seat number'),
            _('Blocked'),
            _('Valid from'),
            _('Valid until'),
            _('Order comment'),
            _('Follow-up date'),
        ]

        questions = list(Question.objects.filter(event__in=self.events))
        options = {}
        for q in questions:
            if q.type == Question.TYPE_CHOICE_MULTIPLE:
                options[q.pk] = []
                if form_data['group_multiple_choice']:
                    for o in q.options.all():
                        options[q.pk].append(o)
                    headers.append(str(q.question))
                else:
                    for o in q.options.all():
                        headers.append(str(q.question) + ' – ' + str(o.answer))
                        options[q.pk].append(o)
            else:
                headers.append(str(q.question))
        headers += [
            _('Company'),
            _('Invoice address name'),
        ]
        if name_scheme and len(name_scheme['fields']) > 1:
            for k, label, w in name_scheme['fields']:
                headers.append(_('Invoice address name') + ': ' + str(label))
        headers += [
            _('Invoice address street'), _('Invoice address ZIP code'), _('Invoice address city'),
            _('Invoice address country'),
            pgettext('address', 'Invoice address state'),
            _('VAT ID'),
        ]
        headers += [
            _('Sales channel'), _('Order locale'),
            _('E-mail address verified'),
            _('External customer ID'),
            _('Payment providers'),
        ]

        # get meta_data labels from first cached event
        meta_data_labels = next(iter(self.event_object_cache.values())).meta_data.keys()
        if has_subevents:
            headers += meta_data_labels
        yield headers

        all_ids = list(base_qs.order_by('order__datetime', 'positionid').values_list('pk', flat=True))
        yield self.ProgressSetTotal(total=len(all_ids))
        for ids in chunked_iterable(all_ids, 1000):
            ops = sorted(qs.filter(id__in=ids), key=lambda k: ids.index(k.pk))

            for op in ops:
                order = op.order
                tz = ZoneInfo(self.event_object_cache[order.event_id].settings.timezone)
                row = [
                    self.event_object_cache[order.event_id].slug,
                    order.code,
                    op.positionid,
                    _("canceled") if op.canceled else order.get_status_display(),
                    order.email,
                    str(order.phone) if order.phone else '',
                    order.datetime.astimezone(tz).strftime('%Y-%m-%d'),
                    order.datetime.astimezone(tz).strftime('%H:%M:%S'),
                ]
                if has_subevents:
                    if op.subevent:
                        row.append(op.subevent.name)
                        row.append(op.subevent.date_from.astimezone(self.event_object_cache[order.event_id].timezone).strftime('%Y-%m-%d %H:%M:%S'))
                        if op.subevent.date_to:
                            row.append(op.subevent.date_to.astimezone(self.event_object_cache[order.event_id].timezone).strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            row.append('')
                    else:
                        row.append('')
                        row.append('')
                        row.append('')
                row += [
                    str(op.item),
                    str(op.variation) if op.variation else '',
                    op.price,
                    op.tax_rate,
                    str(op.tax_rule) if op.tax_rule else '',
                    op.tax_value,
                    op.attendee_name,
                ]
                if name_scheme and len(name_scheme['fields']) > 1:
                    for k, label, w in name_scheme['fields']:
                        row.append(
                            get_name_parts_localized(op.attendee_name_parts, k)
                        )
                row += [
                    op.attendee_email,
                    op.company or '',
                    op.street or '',
                    op.zipcode or '',
                    op.city or '',
                    op.country if op.country else '',
                    op.state or '',
                    op.voucher.code if op.voucher else '',
                    op.pseudonymization_id,
                    op.secret,
                ]

                if op.seat:
                    row += [
                        op.seat.seat_guid,
                        str(op.seat),
                        op.seat.zone_name,
                        op.seat.row_name,
                        op.seat.seat_number,
                    ]
                else:
                    row += ['', '', '', '', '']

                row += [
                    _('Yes') if op.blocked else '',
                    date_format(op.valid_from.astimezone(tz), 'SHORT_DATETIME_FORMAT') if op.valid_from else '',
                    date_format(op.valid_until.astimezone(tz), 'SHORT_DATETIME_FORMAT') if op.valid_until else '',
                ]
                row.append(order.comment)
                row.append(order.custom_followup_at.strftime("%Y-%m-%d") if order.custom_followup_at else "")
                acache = {}
                for a in op.answers.all():
                    # We do not want to localize Date, Time and Datetime question answers, as those can lead
                    # to difficulties parsing the data (for example 2019-02-01 may become Février, 2019 01 in French).
                    if a.question.type == Question.TYPE_CHOICE_MULTIPLE:
                        acache[a.question_id] = set(o.pk for o in a.options.all())
                    elif a.question.type in Question.UNLOCALIZED_TYPES:
                        acache[a.question_id] = a.answer
                    else:
                        acache[a.question_id] = str(a)
                for q in questions:
                    if q.type == Question.TYPE_CHOICE_MULTIPLE:
                        if form_data['group_multiple_choice']:
                            row.append(", ".join(str(o.answer) for o in options[q.pk] if o.pk in acache.get(q.pk, set())))
                        else:
                            for o in options[q.pk]:
                                row.append(_('Yes') if o.pk in acache.get(q.pk, set()) else _('No'))
                    else:
                        row.append(acache.get(q.pk, ''))

                try:
                    row += [
                        order.invoice_address.company,
                        order.invoice_address.name,
                    ]
                    if name_scheme and len(name_scheme['fields']) > 1:
                        for k, label, w in name_scheme['fields']:
                            row.append(
                                get_name_parts_localized(order.invoice_address.name_parts, k)
                            )
                    row += [
                        order.invoice_address.street,
                        order.invoice_address.zipcode,
                        order.invoice_address.city,
                        order.invoice_address.country if order.invoice_address.country else
                        order.invoice_address.country_old,
                        order.invoice_address.state,
                        order.invoice_address.vat_id,
                    ]
                except InvoiceAddress.DoesNotExist:
                    row += [''] * (8 + (len(name_scheme['fields']) if name_scheme and len(name_scheme['fields']) > 1 else 0))
                row += [
                    order.sales_channel,
                    order.locale,
                    _('Yes') if order.email_known_to_work else _('No'),
                    str(order.customer.external_identifier) if order.customer and order.customer.external_identifier else '',
                ]
                row.append(', '.join([
                    str(self.providers.get(p, p)) for p in sorted(set((op.payment_providers or '').split(',')))
                    if p and p != 'free'
                ]))

                if has_subevents:
                    if op.subevent:
                        row += op.subevent.meta_data.values()
                    else:
                        row += [''] * len(meta_data_labels)
                yield row

    def get_filename(self):
        if self.is_multievent:
            return '{}_orders'.format(self.organizer.slug)
        else:
            return '{}_orders'.format(self.event.slug)


class TransactionListExporter(ListExporter):
    identifier = 'transactions'
    verbose_name = gettext_lazy('Order transaction data')
    category = pgettext_lazy('export_category', 'Order data')
    description = gettext_lazy('Download a spreadsheet of all substantial changes to orders, i.e. all changes to '
                               'products, prices or tax rates. The information is only accurate for changes made with '
                               'pretix versions released after October 2021.')

    @cached_property
    def providers(self):
        return dict(get_all_payment_providers())

    @property
    def additional_form_fields(self):
        d = [
            ('date_range',
             DateFrameField(
                 label=_('Date range'),
                 include_future_frames=False,
                 required=False,
                 help_text=_('Only include transactions created within this date range.')
             )),
        ]
        d = OrderedDict(d)
        return d

    @cached_property
    def event_object_cache(self):
        return {e.pk: e for e in self.events}

    def get_filename(self):
        if self.is_multievent:
            return '{}_transactions'.format(self.organizer.slug)
        else:
            return '{}_transactions'.format(self.event.slug)

    def iterate_list(self, form_data):
        qs = Transaction.objects.filter(
            order__event__in=self.events,
        )

        if form_data.get('date_range'):
            dt_start, dt_end = resolve_timeframe_to_datetime_start_inclusive_end_exclusive(now(), form_data['date_range'], self.timezone)
            if dt_start:
                qs = qs.filter(datetime__gte=dt_start)
            if dt_end:
                qs = qs.filter(datetime__lt=dt_end)

        qs = qs.select_related(
            'order', 'order__event', 'item', 'variation', 'subevent',
        ).order_by(
            'datetime', 'id',
        )

        headers = [
            _('Event'),
            _('Event slug'),
            _('Currency'),

            _('Order code'),
            _('Order date'),
            _('Order time'),

            _('Transaction date'),
            _('Transaction time'),
            _('Old data'),

            _('Position ID'),
            _('Quantity'),

            _('Product ID'),
            _('Product'),
            _('Variation ID'),
            _('Variation'),
            _('Fee type'),
            _('Internal fee type'),

            pgettext('subevent', 'Date ID'),
            pgettext('subevent', 'Date'),

            _('Price'),
            _('Tax rate'),
            _('Tax rule ID'),
            _('Tax rule'),
            _('Tax value'),
            _('Gross total'),
            _('Tax total'),
        ]

        if form_data.get('_format') == 'xlsx':
            for i in range(len(headers)):
                headers[i] = WriteOnlyCell(self.__ws, value=headers[i])
                if i in (0, 12, 14, 18, 22):
                    headers[i].fill = PatternFill(start_color="FFB419", end_color="FFB419", fill_type="solid")
                    headers[i].comment = Comment(
                        text=_(
                            "This value is supplied for informational purposes, it is not part of the original transaction "
                            "data and might have changed since the transaction."
                        ),
                        author='system'
                    )
                headers[i].font = Font(bold=True)

        yield headers

        yield self.ProgressSetTotal(total=qs.count())

        for t in qs.iterator():
            row = [
                str(t.order.event.name),
                t.order.event.slug,
                t.order.event.currency,

                t.order.code,
                t.order.datetime.astimezone(self.timezone).strftime('%Y-%m-%d'),
                t.order.datetime.astimezone(self.timezone).strftime('%H:%M:%S'),

                t.datetime.astimezone(self.timezone).strftime('%Y-%m-%d'),
                t.datetime.astimezone(self.timezone).strftime('%H:%M:%S'),
                _('Converted from legacy version') if t.migrated else '',

                t.positionid,
                t.count,

                t.item_id,
                str(t.item),
                t.variation_id or '',
                str(t.variation) if t.variation_id else '',
                t.fee_type,
                t.internal_type,
                t.subevent_id or '',
                str(t.subevent) if t.subevent else '',

                t.price,
                t.tax_rate,
                t.tax_rule_id or '',
                str(t.tax_rule.internal_name or t.tax_rule.name) if t.tax_rule_id else '',
                t.tax_value,
                t.price * t.count,
                t.tax_value * t.count,
            ]

            if form_data.get('_format') == 'xlsx':
                for i in range(len(row)):
                    if t.order.testmode:
                        row[i] = WriteOnlyCell(self.__ws, value=remove_invalid_excel_chars(row[i]))
                        row[i].fill = PatternFill(start_color="FFB419", end_color="FFB419", fill_type="solid")

            yield row

    def prepare_xlsx_sheet(self, ws):
        self.__ws = ws
        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 25
        ws.column_dimensions['N'].width = 10
        ws.column_dimensions['O'].width = 25
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 20
        ws.column_dimensions['R'].width = 10
        ws.column_dimensions['S'].width = 25
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 10
        ws.column_dimensions['W'].width = 20
        ws.column_dimensions['X'].width = 15


class PaymentListExporter(ListExporter):
    identifier = 'paymentlist'
    verbose_name = gettext_lazy('Payments and refunds')
    category = pgettext_lazy('export_category', 'Order data')
    description = gettext_lazy('Download a spreadsheet of all payments or refunds of every order.')
    featured = True

    @property
    def additional_form_fields(self):
        return OrderedDict(
            [
                ('end_date_range',
                 DateFrameField(
                     label=_('Date range (payment date)'),
                     include_future_frames=False,
                     required=False,
                     help_text=_('Note that using this will exclude any non-confirmed payments or non-completed refunds.'),
                 )),
                ('start_date_range',
                 DateFrameField(
                     label=_('Date range (start of transaction)'),
                     include_future_frames=False,
                     required=False
                 )),
                ('payment_states',
                 forms.MultipleChoiceField(
                     label=_('Payment states'),
                     choices=OrderPayment.PAYMENT_STATES,
                     initial=[OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED],
                     required=False,
                     widget=forms.CheckboxSelectMultiple,
                 )),
                ('refund_states',
                 forms.MultipleChoiceField(
                     label=_('Refund states'),
                     choices=OrderRefund.REFUND_STATES,
                     initial=[OrderRefund.REFUND_STATE_DONE, OrderRefund.REFUND_STATE_CREATED,
                              OrderRefund.REFUND_STATE_TRANSIT],
                     widget=forms.CheckboxSelectMultiple,
                     required=False
                 )),
            ]
        )

    def iterate_list(self, form_data):
        provider_names = dict(get_all_payment_providers())

        payments = OrderPayment.objects.filter(
            order__event__in=self.events,
            state__in=form_data.get('payment_states', [])
        ).select_related('order').prefetch_related('order__event').order_by('created')
        refunds = OrderRefund.objects.filter(
            order__event__in=self.events,
            state__in=form_data.get('refund_states', [])
        ).select_related('order').prefetch_related('order__event').order_by('created')

        if form_data.get('end_date_range'):
            dt_start, dt_end = resolve_timeframe_to_datetime_start_inclusive_end_exclusive(now(), form_data['end_date_range'], self.timezone)
            if dt_start:
                payments = payments.filter(created__gte=dt_start)
                refunds = refunds .filter(created__gte=dt_start)
            if dt_end:
                payments = payments.filter(created__lt=dt_end)
                refunds = refunds .filter(created__lt=dt_end)

        if form_data.get('start_end_date_range'):
            dt_start, dt_end = resolve_timeframe_to_datetime_start_inclusive_end_exclusive(now(), form_data['start_date_range'], self.timezone)
            if dt_start:
                payments = payments.filter(payment_date__gte=dt_start)
                refunds = refunds .filter(execution_date__gte=dt_start)
            if dt_end:
                payments = payments.filter(payment_date__lt=dt_end)
                refunds = refunds.filter(execution_date__lt=dt_end)

        objs = sorted(list(payments) + list(refunds), key=lambda o: o.created)

        headers = [
            _('Event slug'), _('Order'), _('Payment ID'), _('Creation date'), _('Completion date'), _('Status'),
            _('Status code'), _('Amount'), _('Payment method'), _('Comment'), _('Matching ID'), _('Payment details'),
        ]
        yield headers

        yield self.ProgressSetTotal(total=len(objs))
        for obj in objs:
            tz = ZoneInfo(obj.order.event.settings.timezone)
            if isinstance(obj, OrderPayment) and obj.payment_date:
                d2 = obj.payment_date.astimezone(tz).date().strftime('%Y-%m-%d')
            elif isinstance(obj, OrderRefund) and obj.execution_date:
                d2 = obj.execution_date.astimezone(tz).date().strftime('%Y-%m-%d')
            else:
                d2 = ''
            matching_id = ''
            payment_details = ''
            try:
                if isinstance(obj, OrderPayment):
                    matching_id = obj.payment_provider.matching_id(obj) or ''
                    payment_details = obj.payment_provider.payment_control_render_short(obj)
                elif isinstance(obj, OrderRefund):
                    matching_id = obj.payment_provider.refund_matching_id(obj) or ''
                    payment_details = obj.payment_provider.refund_control_render_short(obj)
            except Exception:
                pass

            row = [
                obj.order.event.slug,
                obj.order.code,
                obj.full_id,
                obj.created.astimezone(tz).date().strftime('%Y-%m-%d'),
                d2,
                obj.get_state_display(),
                obj.state,
                obj.amount * (-1 if isinstance(obj, OrderRefund) else 1),
                provider_names.get(obj.provider, obj.provider),
                obj.comment if isinstance(obj, OrderRefund) else "",
                matching_id,
                payment_details,
            ]
            yield row

    def get_filename(self):
        if self.is_multievent:
            return '{}_payments'.format(self.organizer.slug)
        else:
            return '{}_payments'.format(self.event.slug)


class QuotaListExporter(ListExporter):
    identifier = 'quotalist'
    verbose_name = gettext_lazy('Quota availabilities')
    category = pgettext_lazy('export_category', 'Product data')
    description = gettext_lazy('Download a spreadsheet of all quotas including their current availability.')

    def iterate_list(self, form_data):
        has_subevents = self.event.has_subevents
        headers = [
            _('Quota name'), _('Total quota'), _('Paid orders'), _('Pending orders'), _('Blocking vouchers'),
            _('Current user\'s carts'), _('Waiting list'), _('Exited orders'), _('Current availability')
        ]
        if has_subevents:
            headers.append(pgettext('subevent', 'Date'))
            headers.append(_('Start date'))
            headers.append(_('End date'))
        yield headers

        quotas = list(self.event.quotas.select_related('subevent'))
        qa = QuotaAvailability(full_results=True)
        qa.queue(*quotas)
        qa.compute()

        for quota in quotas:
            avail = qa.results[quota]
            row = [
                quota.name,
                _('Infinite') if quota.size is None else quota.size,
                qa.count_paid_orders[quota],
                qa.count_pending_orders[quota],
                qa.count_vouchers[quota],
                qa.count_cart[quota],
                qa.count_waitinglist[quota],
                qa.count_exited_orders[quota],
                _('Infinite') if avail[1] is None else avail[1]
            ]
            if has_subevents:
                if quota.subevent:
                    row.append(quota.subevent.name)
                    row.append(quota.subevent.date_from.astimezone(self.event.timezone).strftime('%Y-%m-%d %H:%M:%S'))
                    if quota.subevent.date_to:
                        row.append(quota.subevent.date_to.astimezone(self.event.timezone).strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        row.append('')
                else:
                    row.append('')
                    row.append('')
                    row.append('')
            yield row

    def get_filename(self):
        return '{}_quotas'.format(self.event.slug)


class GiftcardTransactionListExporter(OrganizerLevelExportMixin, ListExporter):
    identifier = 'giftcardtransactionlist'
    verbose_name = gettext_lazy('Gift card transactions')
    organizer_required_permission = 'can_manage_gift_cards'
    category = pgettext_lazy('export_category', 'Gift cards')
    description = gettext_lazy('Download a spreadsheet of all gift card transactions.')

    @property
    def additional_form_fields(self):
        d = [
            ('date_range',
             DateFrameField(
                 label=_('Date range'),
                 include_future_frames=False,
                 required=False
             )),
        ]
        d = OrderedDict(d)
        return d

    def iterate_list(self, form_data):
        qs = GiftCardTransaction.objects.filter(
            card__issuer=self.organizer,
        ).order_by('datetime').select_related('card', 'order', 'order__event', 'acceptor')

        if form_data.get('date_range'):
            dt_start, dt_end = resolve_timeframe_to_datetime_start_inclusive_end_exclusive(now(), form_data['date_range'], self.timezone)
            if dt_start:
                qs = qs.filter(datetime__gte=dt_start)
            if dt_end:
                qs = qs.filter(datetime__lt=dt_end)

        headers = [
            _('Gift card code'),
            _('Test mode'),
            _('Date'),
            _('Amount'),
            _('Currency'),
            _('Order'),
            _('Organizer'),
        ]
        yield headers

        for obj in qs:
            row = [
                obj.card.secret,
                _('TEST MODE') if obj.card.testmode else '',
                obj.datetime.astimezone(self.timezone).strftime('%Y-%m-%d %H:%M:%S'),
                obj.value,
                obj.card.currency,
                obj.order.full_code if obj.order else None,
                str(obj.acceptor or ""),
            ]
            yield row

    def get_filename(self):
        return '{}_giftcardtransactions'.format(self.organizer.slug)


class GiftcardRedemptionListExporter(ListExporter):
    identifier = 'giftcardredemptionlist'
    verbose_name = gettext_lazy('Gift card redemptions')
    category = pgettext_lazy('export_category', 'Order data')
    description = gettext_lazy('Download a spreadsheet of all payments or refunds that involve gift cards.')

    def iterate_list(self, form_data):
        payments = OrderPayment.objects.filter(
            order__event__in=self.events,
            provider='giftcard',
            state__in=(OrderPayment.PAYMENT_STATE_CONFIRMED, OrderPayment.PAYMENT_STATE_REFUNDED),
        ).order_by('created')
        refunds = OrderRefund.objects.filter(
            order__event__in=self.events,
            provider='giftcard',
            state=OrderRefund.REFUND_STATE_DONE
        ).order_by('created')

        objs = sorted(list(payments) + list(refunds), key=lambda o: (o.order.code, o.created))

        headers = [
            _('Event slug'), _('Order'), _('Payment ID'), _('Date'), _('Gift card code'), _('Amount'), _('Issuer')
        ]
        yield headers

        for obj in objs:
            tz = ZoneInfo(obj.order.event.settings.timezone)
            gc = GiftCard.objects.get(pk=obj.info_data.get('gift_card'))
            row = [
                obj.order.event.slug,
                obj.order.code,
                obj.full_id,
                obj.created.astimezone(tz).date().strftime('%Y-%m-%d'),
                gc.secret,
                obj.amount * (-1 if isinstance(obj, OrderRefund) else 1),
                gc.issuer
            ]
            yield row

    def get_filename(self):
        if self.is_multievent:
            return '{}_giftcardredemptions'.format(self.organizer.slug)
        else:
            return '{}_giftcardredemptions'.format(self.event.slug)


class GiftcardListExporter(OrganizerLevelExportMixin, ListExporter):
    identifier = 'giftcardlist'
    verbose_name = gettext_lazy('Gift cards')
    organizer_required_permission = 'can_manage_gift_cards'
    category = pgettext_lazy('export_category', 'Gift cards')
    description = gettext_lazy('Download a spreadsheet of all gift cards including their current value.')

    @property
    def additional_form_fields(self):
        return OrderedDict(
            [
                ('date', forms.SplitDateTimeField(
                    label=_('Show value at'),
                    required=False,
                    widget=SplitDateTimePickerWidget(),
                    help_text=_('Defaults to the time of report.')
                )),
                ('testmode', forms.ChoiceField(
                    label=_('Test mode'),
                    choices=(
                        ('', _('All')),
                        ('yes', _('Test mode')),
                        ('no', _('Live')),
                    ),
                    initial='no',
                    required=False
                )),
                ('state', forms.ChoiceField(
                    label=_('Status'),
                    choices=(
                        ('', _('All')),
                        ('empty', _('Empty')),
                        ('valid_value', _('Valid and with value')),
                        ('expired_value', _('Expired and with value')),
                        ('expired', _('Expired')),
                    ),
                    initial='valid_value',
                    required=False
                ))
            ]
        )

    def iterate_list(self, form_data):
        d = form_data.get('date') or now()
        s = GiftCardTransaction.objects.filter(
            card=OuterRef('pk'),
            datetime__lte=d
        ).order_by().values('card').annotate(s=Sum('value')).values('s')
        qs = self.organizer.issued_gift_cards.filter(
            issuance__lte=d
        ).annotate(
            cached_value=Coalesce(Subquery(s), Decimal('0.00')),
        ).order_by('issuance').prefetch_related(
            'transactions', 'transactions__order', 'transactions__order__event', 'transactions__order__invoices'
        )

        if form_data.get('testmode') == 'yes':
            qs = qs.filter(testmode=True)
        elif form_data.get('testmode') == 'no':
            qs = qs.filter(testmode=False)

        if form_data.get('state') == 'empty':
            qs = qs.filter(cached_value=0)
        elif form_data.get('state') == 'valid_value':
            qs = qs.exclude(cached_value=0).filter(Q(expires__isnull=True) | Q(expires__gte=d))
        elif form_data.get('state') == 'expired_value':
            qs = qs.exclude(cached_value=0).filter(expires__lt=d)
        elif form_data.get('state') == 'expired':
            qs = qs.filter(expires__lt=d)

        headers = [
            _('Gift card code'),
            _('Test mode card'),
            _('Creation date'),
            _('Expiry date'),
            _('Special terms and conditions'),
            _('Currency'),
            _('Current value'),
            _('Created in order'),
            _('Last invoice number of order'),
            _('Last invoice date of order'),
        ]
        yield headers

        tz = get_current_timezone()
        for obj in qs:
            o = None
            i = None
            trans = list(obj.transactions.all())
            if trans:
                o = trans[0].order
            if o:
                invs = list(o.invoices.all())
                if invs:
                    i = invs[-1]
            row = [
                obj.secret,
                _('Yes') if obj.testmode else _('No'),
                obj.issuance.astimezone(tz).date().strftime('%Y-%m-%d'),
                obj.expires.astimezone(tz).date().strftime('%Y-%m-%d') if obj.expires else '',
                obj.conditions or '',
                obj.currency,
                obj.cached_value,
                o.full_code if o else '',
                i.number if i else '',
                i.date.strftime('%Y-%m-%d') if i else '',
            ]
            yield row

    def get_filename(self):
        return '{}_giftcards'.format(self.organizer.slug)


@receiver(register_data_exporters, dispatch_uid="exporter_orderlist")
def register_orderlist_exporter(sender, **kwargs):
    return OrderListExporter


@receiver(register_multievent_data_exporters, dispatch_uid="multiexporter_orderlist")
def register_multievent_orderlist_exporter(sender, **kwargs):
    return OrderListExporter


@receiver(register_data_exporters, dispatch_uid="exporter_ordertransactionlist")
def register_ordertransactionlist_exporter(sender, **kwargs):
    return TransactionListExporter


@receiver(register_multievent_data_exporters, dispatch_uid="multiexporter_ordertransactionlist")
def register_multievent_ordertransactionlist_exporter(sender, **kwargs):
    return TransactionListExporter


@receiver(register_data_exporters, dispatch_uid="exporter_paymentlist")
def register_paymentlist_exporter(sender, **kwargs):
    return PaymentListExporter


@receiver(register_multievent_data_exporters, dispatch_uid="multiexporter_paymentlist")
def register_multievent_paymentlist_exporter(sender, **kwargs):
    return PaymentListExporter


@receiver(register_data_exporters, dispatch_uid="exporter_quotalist")
def register_quotalist_exporter(sender, **kwargs):
    return QuotaListExporter


@receiver(register_data_exporters, dispatch_uid="exporter_giftcardredemptionlist")
def register_giftcardredemptionlist_exporter(sender, **kwargs):
    return GiftcardRedemptionListExporter


@receiver(register_multievent_data_exporters, dispatch_uid="multiexporter_giftcardredemptionlist")
def register_multievent_i_giftcardredemptionlist_exporter(sender, **kwargs):
    return GiftcardRedemptionListExporter


@receiver(register_multievent_data_exporters, dispatch_uid="multiexporter_giftcardlist")
def register_multievent_i_giftcardlist_exporter(sender, **kwargs):
    return GiftcardListExporter


@receiver(register_multievent_data_exporters, dispatch_uid="multiexporter_giftcardtransactionlist")
def register_multievent_i_giftcardtransactionlist_exporter(sender, **kwargs):
    return GiftcardTransactionListExporter
