#
# This file is part of pretix (Community Edition).
#
# Copyright (C) 2014-2020 Raphael Michel and contributors
# Copyright (C) 2020-2021 rami.io GmbH and contributors
#
# This program is free software: you can redistribute it and/or modify it under the terms of the GNU Affero General
# Public License as published by the Free Software Foundation in version 3 of the License.
#
# ADDITIONAL TERMS APPLY: Pursuant to Section 7 of the GNU Affero General Public License, additional terms are
# applicable granting you additional permissions and placing additional restrictions on your usage of this software.
# Please refer to the pretix LICENSE file to obtain the full terms applicable to this work. If you did not receive
# this file, see <https://pretix.eu/about/en/license>.
#
# This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even the implied
# warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU Affero General Public License for more
# details.
#
# You should have received a copy of the GNU Affero General Public License along with this program.  If not, see
# <https://www.gnu.org/licenses/>.
#
from django.db.models import Prefetch
from django.dispatch import receiver
from django.utils.formats import date_format
from django.utils.translation import gettext_lazy as _, pgettext_lazy
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from ...helpers.safe_openpyxl import SafeCell
from ..channels import get_all_sales_channels
from ..exporter import ListExporter
from ..models import ItemMetaValue, ItemVariation, ItemVariationMetaValue
from ..signals import register_data_exporters


def _max(a1, a2):
    if a1 and a2:
        return max(a1, a2)
    return a1 or a2


def _min(a1, a2):
    if a1 and a2:
        return min(a1, a2)
    return a1 or a2


class ItemDataExporter(ListExporter):
    identifier = 'itemdata'
    verbose_name = _('Product data')
    category = pgettext_lazy('export_category', 'Product data')
    description = _('Download a spreadsheet with details about all products and variations.')

    def iterate_list(self, form_data):
        locales = self.event.settings.locales
        scs = get_all_sales_channels()
        header = [
            _("Product ID"),
            _("Variation ID"),
            _("Product category"),
            _("Internal name"),
        ]
        for l in locales:
            header.append(
                _("Item name") + f" ({l})"
            )
        for l in locales:
            header.append(
                _("Variation") + f" ({l})"
            )
        header += [
            _("Active"),
            _("Sales channels"),
            _("Default price"),
            _("Free price input"),
            _("Sales tax"),
            _("Is an admission ticket"),
            _("Personalized ticket"),
            _("Generate tickets"),
            _("Waiting list"),
            _("Available from"),
            _("Available until"),
            _("This product can only be bought using a voucher."),
            _("This product will only be shown if a voucher matching the product is redeemed."),
            _("Buying this product requires approval"),
            _("Only sell this product as part of a bundle"),
            _("Allow product to be canceled or changed"),
            _("Minimum amount per order"),
            _("Maximum amount per order"),
            _("Requires special attention"),
            _("Original price"),
            _("This product is a gift card"),
            _("Require a valid membership"),
            _("Hide without a valid membership"),
        ]
        props = list(self.event.item_meta_properties.all())
        for p in props:
            header.append(p.name)

        if form_data["_format"] == "xlsx":
            row = []
            for h in header:
                c = SafeCell(self.__ws, value=h)
                c.alignment = Alignment(wrap_text=True, vertical='top')
                row.append(c)
        else:
            row = header

        yield row

        for i in self.event.items.prefetch_related(
            Prefetch(
                'meta_values',
                ItemMetaValue.objects.select_related('property'),
                to_attr='meta_values_cached'
            ),
            Prefetch(
                'variations',
                queryset=ItemVariation.objects.prefetch_related(
                    Prefetch(
                        'meta_values',
                        ItemVariationMetaValue.objects.select_related('property'),
                        to_attr='meta_values_cached'
                    ),
                ),
            ),
        ).select_related('category', 'tax_rule'):
            vars = list(i.variations.all())

            if vars:
                for v in vars:
                    m = v.meta_data
                    row = [
                        i.pk,
                        v.pk,
                        str(i.category) if i.category else "",
                        i.internal_name or "",
                    ]
                    for l in locales:
                        row.append(i.name.localize(l))
                    for l in locales:
                        row.append(v.value.localize(l))
                    row += [
                        _("Yes") if i.active and v.active else "",
                        ", ".join([str(sn.verbose_name) for s, sn in scs.items() if s in i.sales_channels and s in v.sales_channels]),
                        v.default_price or i.default_price,
                        _("Yes") if i.free_price else "",
                        str(i.tax_rule) if i.tax_rule else "",
                        _("Yes") if i.admission else "",
                        _("Yes") if i.personalized else "",
                        _("Yes") if i.generate_tickets else (_("Default") if i.generate_tickets is None else ""),
                        _("Yes") if i.allow_waitinglist else "",
                        date_format(_max(i.available_from, v.available_from).astimezone(self.timezone),
                                    "SHORT_DATETIME_FORMAT") if i.available_from or v.available_from else "",
                        date_format(_min(i.available_until, v.available_until).astimezone(self.timezone),
                                    "SHORT_DATETIME_FORMAT") if i.available_until or v.available_until else "",
                        _("Yes") if i.require_voucher else "",
                        _("Yes") if i.hide_without_voucher or v.hide_without_voucher else "",
                        _("Yes") if i.require_approval or v.require_approval else "",
                        _("Yes") if i.require_bundling else "",
                        _("Yes") if i.allow_cancel else "",
                        i.min_per_order if i.min_per_order is not None else "",
                        i.max_per_order if i.max_per_order is not None else "",
                        _("Yes") if i.checkin_attention else "",
                        v.original_price or i.original_price or "",
                        _("Yes") if i.issue_giftcard else "",
                        _("Yes") if i.require_membership or v.require_membership else "",
                        _("Yes") if i.require_membership_hidden or v.require_membership_hidden else "",
                    ]
                    row += [
                        m.get(p.name, '') for p in props
                    ]
                    yield row

            else:
                m = i.meta_data
                row = [
                    i.pk,
                    "",
                    str(i.category) if i.category else "",
                    i.internal_name or "",
                ]
                for l in locales:
                    row.append(i.name.localize(l))
                for l in locales:
                    row.append("")
                row += [
                    _("Yes") if i.active else "",
                    ", ".join([str(sn.verbose_name) for s, sn in scs.items() if s in i.sales_channels]),
                    i.default_price,
                    _("Yes") if i.free_price else "",
                    str(i.tax_rule) if i.tax_rule else "",
                    _("Yes") if i.admission else "",
                    _("Yes") if i.personalized else "",
                    _("Yes") if i.generate_tickets else (_("Default") if i.generate_tickets is None else ""),
                    _("Yes") if i.allow_waitinglist else "",
                    date_format(i.available_from.astimezone(self.timezone),
                                "SHORT_DATETIME_FORMAT") if i.available_from else "",
                    date_format(i.available_until.astimezone(self.timezone),
                                "SHORT_DATETIME_FORMAT") if i.available_until else "",
                    _("Yes") if i.require_voucher else "",
                    _("Yes") if i.hide_without_voucher else "",
                    _("Yes") if i.require_approval else "",
                    _("Yes") if i.require_bundling else "",
                    _("Yes") if i.allow_cancel else "",
                    i.min_per_order if i.min_per_order is not None else "",
                    i.max_per_order if i.max_per_order is not None else "",
                    _("Yes") if i.checkin_attention else "",
                    i.original_price or "",
                    _("Yes") if i.issue_giftcard else "",
                    _("Yes") if i.require_membership else "",
                    _("Yes") if i.require_membership_hidden else "",
                ]

                row += [
                    m.get(p.name, '') for p in props
                ]
                yield row

    def get_filename(self):
        if self.is_multievent:
            return '{}_products'.format(self.organizer.slug)
        return '{}_products'.format(self.event.slug)

    def prepare_xlsx_sheet(self, ws):
        self.__ws = ws
        ws.freeze_panes = 'A1'
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        for i in range(len(self.event.settings.locales)):
            ws.column_dimensions[get_column_letter(5 + 2 * i + 0)].width = 25
            ws.column_dimensions[get_column_letter(5 + 2 * i + 1)].width = 25
        ws.column_dimensions[get_column_letter(5 + 2 * len(self.event.settings.locales) + 1)].width = 25
        ws.row_dimensions[1].height = 40


@receiver(register_data_exporters, dispatch_uid="exporter_itemdata")
def register_itemdata_exporter(sender, **kwargs):
    return ItemDataExporter
