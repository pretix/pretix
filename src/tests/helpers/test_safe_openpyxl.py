#
# This file is part of pretix (Community Edition).
#
# Copyright (C) 2014-2020 Raphael Michel and contributors
# Copyright (C) 2020-2021 rami.io GmbH and contributors
#
# This program is free software: you can redistribute it and/or modify it under the terms of the GNU Affero General
# Public License as published by the Free Software Foundation in version 3 of the License.
#
# ADDITIONAL TERMS APPLY: Pursuant to Section 7 of the GNU Affero General Public License, additional terms are
# applicable granting you additional permissions and placing additional restrictions on your usage of this software.
# Please refer to the pretix LICENSE file to obtain the full terms applicable to this work. If you did not receive
# this file, see <https://pretix.eu/about/en/license>.
#
# This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even the implied
# warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU Affero General Public License for more
# details.
#
# You should have received a copy of the GNU Affero General Public License along with this program.  If not, see
# <https://www.gnu.org/licenses/>.
#
from openpyxl.cell.cell import TYPE_STRING

from pretix.helpers.safe_openpyxl import SafeWorkbook


def test_invalid_byte_removed():
    wb = SafeWorkbook()
    ws = wb.create_sheet()
    ws.append(["foo\u0000bar"])
    assert ws.cell(1, 1).value == "foobar"
    ws.append(["foo\uffffbaz"])
    assert ws.cell(2, 1).value == "foobaz"


def test_no_formulas():
    wb = SafeWorkbook()
    ws = wb.create_sheet()
    ws.append(["=1+1"])
    assert ws.cell(1, 1).data_type == TYPE_STRING
